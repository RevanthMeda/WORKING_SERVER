import csv
import io
import logging
import os
import re
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    from openpyxl import load_workbook  # type: ignore
except ImportError:  # pragma: no cover - handled via warnings at runtime
    load_workbook = None

try:
    from docx import Document  # type: ignore
except ImportError:  # pragma: no cover - optional dependency
    Document = None

try:
    import fitz  # type: ignore
except ImportError:  # pragma: no cover - optional dependency
    fitz = None

from services.sat_tables import TABLE_CONFIG

logger = logging.getLogger(__name__)

_NORMALIZE = re.compile(r'[^a-z0-9]+')


@dataclass
class AutoFillResult:
    """Container for auto-fill results returned to the chatbot assistant."""

    field_updates: Dict[str, str] = field(default_factory=dict)
    table_updates: Dict[str, List[Dict[str, str]]] = field(default_factory=dict)
    messages: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)

    def extend(self, other: "AutoFillResult") -> None:
        """Merge another result into this one."""
        if other.field_updates:
            self.field_updates.update(other.field_updates)
        if other.table_updates:
            for key, rows in other.table_updates.items():
                self.table_updates.setdefault(key, []).extend(rows)
        self.messages.extend(other.messages)
        self.warnings.extend(other.warnings)


class SatFormAutoFillParser:
    """Parse uploaded artefacts and map them to SAT fields and table sections."""

    TABULAR_EXTENSIONS = {'.csv', '.tsv', '.xlsx', '.xlsm', '.xls'}
    TEXT_EXTENSIONS = {'.txt', '.md'}
    DOCX_EXTENSIONS = {'.docx'}
    PDF_EXTENSIONS = {'.pdf'}

    def __init__(self) -> None:
        self.section_index = self._build_section_index()

    def process(self, file_storage) -> AutoFillResult:
        """Process a Werkzeug FileStorage upload."""
        filename = (file_storage.filename or '').strip() or 'uploaded_file'
        extension = os.path.splitext(filename)[1].lower()

        if extension in self.TABULAR_EXTENSIONS:
            return self._process_tabular(file_storage, filename, extension)
        if extension in self.TEXT_EXTENSIONS or extension in self.DOCX_EXTENSIONS or extension in self.PDF_EXTENSIONS:
            return self._process_textual(file_storage, filename, extension)

        result = AutoFillResult()
        result.warnings.append(f"Auto-fill is not configured for {filename} ({extension or 'unknown format'}).")
        return result

    # ------------------------------------------------------------------ #
    # Tabular ingestion
    # ------------------------------------------------------------------ #
    def _process_tabular(self, file_storage, filename: str, extension: str) -> AutoFillResult:
        datasets = []

        try:
            if extension == '.csv' or extension == '.tsv':
                dataset = self._read_csv_dataset(file_storage, filename, extension)
                if dataset:
                    datasets.append(dataset)
            else:
                if load_workbook is None:
                    warning = "Install openpyxl to enable Excel auto-fill ingestion."
                    logger.warning(warning)
                    return AutoFillResult(warnings=[warning])
                datasets.extend(self._read_excel_datasets(file_storage, filename))
        except Exception as exc:  # noqa: BLE001 - user feedback preferred
            logger.exception("Failed to process tabular upload")
            return AutoFillResult(warnings=[f"Unable to read {filename}: {exc}"])

        result = AutoFillResult()
        for source_label, headers, rows in datasets:
            if not headers or not rows:
                continue
            header_map = self._match_section(headers)
            if not header_map:
                continue
            mapped_rows = self._map_rows(rows, header_map)
            if not mapped_rows:
                continue
            section_key = header_map["section"]
            result.table_updates.setdefault(section_key, []).extend(mapped_rows)
            pretty = self._humanize_section(section_key)
            result.messages.append(
                f"Auto-filled {len(mapped_rows)} row(s) for {pretty} from {source_label}."
            )

        if not result.table_updates:
            result.warnings.append(f"Could not identify any SAT table sections inside {filename}.")
        return result

    def _read_csv_dataset(self, file_storage, filename: str, extension: str) -> Optional[Tuple[str, List[str], List[Dict[str, str]]]]:
        delimiter = '\t' if extension == '.tsv' else ','
        content = file_storage.read()
        file_storage.stream.seek(0)
        if not content:
            return None
        try:
            text = content.decode('utf-8-sig')
        except UnicodeDecodeError:
            text = content.decode('latin-1', errors='ignore')
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        headers = reader.fieldnames or []
        rows = [self._clean_row(row) for row in reader if row]
        return (filename, headers, rows)

    def _read_excel_datasets(self, file_storage, filename: str) -> List[Tuple[str, List[str], List[Dict[str, str]]]]:
        datasets: List[Tuple[str, List[str], List[Dict[str, str]]]] = []
        workbook = load_workbook(file_storage, data_only=True, read_only=True)  # type: ignore[arg-type]
        try:
            for sheet in workbook.worksheets:
                headers, rows = self._sheet_to_rows(sheet)
                if headers and rows:
                    label = f"{filename}:{sheet.title}"
                    datasets.append((label, headers, rows))
        finally:
            file_storage.stream.seek(0)
            workbook.close()
        return datasets

    def _sheet_to_rows(self, sheet) -> Tuple[List[str], List[Dict[str, str]]]:
        headers: List[str] = []
        rows: List[Dict[str, str]] = []
        header_found = False
        blank_streak = 0

        for row in sheet.iter_rows(values_only=True):
            normalized = [self._coerce(value) for value in row]
            if not header_found:
                non_empty_headers = [cell for cell in normalized if cell]
                if len(non_empty_headers) >= 2:
                    headers = [cell or f"Column{idx+1}" for idx, cell in enumerate(normalized)]
                    header_found = True
                continue

            if not any(normalized):
                blank_streak += 1
                if blank_streak >= 20:
                    break
                continue

            blank_streak = 0
            row_dict = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                value = normalized[idx] if idx < len(normalized) else ""
                row_dict[header] = value
            if any(value for value in row_dict.values()):
                rows.append(row_dict)
        return headers, rows

    def _match_section(self, headers: Iterable[str]) -> Optional[Dict[str, Any]]:
        best_match: Optional[Dict[str, Any]] = None
        header_list = [header for header in headers if header]
        if not header_list:
            return None

        for section_name, section_meta in self.section_index.items():
            matches = {}
            unique_fields = set()
            total_score = 0.0

            for header in header_list:
                normalized = self._normalize(header)
                if not normalized:
                    continue
                field_info = self._best_field_match(normalized, section_meta["fields"])
                if not field_info:
                    continue
                matches[header] = field_info
                unique_fields.add(field_info["ui"])
                total_score += field_info["score"]

            if len(unique_fields) < section_meta["min_required"]:
                continue

            coverage = len(unique_fields) / max(len(section_meta["fields"]), 1)
            confidence = total_score / max(len(matches), 1)
            ranking = (len(unique_fields), coverage, confidence)
            if not best_match or ranking > best_match["ranking"]:
                best_match = {
                    "section": section_name,
                    "matches": matches,
                    "ranking": ranking,
                }

        return best_match

    def _map_rows(self, rows: List[Dict[str, str]], header_map: Dict[str, Any]) -> List[Dict[str, str]]:
        mapped: List[Dict[str, str]] = []
        matches = header_map["matches"]
        for row in rows:
            mapped_row: Dict[str, str] = {}
            for header, field_info in matches.items():
                raw_value = row.get(header)
                if not raw_value:
                    continue
                mapped_row[field_info["ui"]] = raw_value
            if mapped_row:
                mapped.append(mapped_row)
        return mapped

    def _best_field_match(self, value: str, fields: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
        best_field = None
        best_score = 0.0
        for field in fields:
            for alias in field["aliases"]:
                score = self._similarity(value, alias)
                if score > best_score:
                    best_score = score
                    best_field = field
            if best_score >= 0.95:  # Early exit on very strong match
                break
        if best_field and best_score >= 0.58:
            return {**best_field, "score": best_score}
        return None

    # ------------------------------------------------------------------ #
    # Textual ingestion (scope/purpose narratives)
    # ------------------------------------------------------------------ #
    def _process_textual(self, file_storage, filename: str, extension: str) -> AutoFillResult:
        text, warning = self._extract_text(file_storage, extension)
        if warning:
            return AutoFillResult(warnings=[warning])
        if not text:
            return AutoFillResult(warnings=[f"{filename} did not contain readable text."])

        target_field = self._infer_text_field(filename, text)
        if not target_field:
            return AutoFillResult(warnings=[f"Could not determine which field to fill from {filename}."])

        cleaned = self._normalize_text_block(text)
        if not cleaned:
            return AutoFillResult(warnings=[f"{filename} contained only empty text."])

        label = target_field.replace('_', ' ').title()
        return AutoFillResult(
            field_updates={target_field: cleaned},
            messages=[f"Filled {label} using {filename}."]
        )

    def _extract_text(self, file_storage, extension: str) -> Tuple[str, Optional[str]]:
        if extension in self.TEXT_EXTENSIONS:
            content = file_storage.read()
            file_storage.stream.seek(0)
            text = content.decode('utf-8-sig', errors='ignore')
            return text, None

        if extension in self.DOCX_EXTENSIONS:
            if Document is None:
                warning = "Install python-docx to enable DOCX ingestion."
                logger.warning(warning)
                return "", warning
            data = file_storage.read()
            file_storage.stream.seek(0)
            document = Document(io.BytesIO(data))
            paragraphs = [para.text for para in document.paragraphs if para.text]
            return "\n".join(paragraphs), None

        if extension in self.PDF_EXTENSIONS:
            if fitz is None:
                warning = "Install PyMuPDF to enable PDF ingestion."
                logger.warning(warning)
                return "", warning
            data = file_storage.read()
            file_storage.stream.seek(0)
            text_fragments = []
            with fitz.open(stream=data, filetype="pdf") as pdf_doc:  # type: ignore[attr-defined]
                for page in pdf_doc:  # pragma: no branch - depends on PDF length
                    text_fragments.append(page.get_text())
            return "\n".join(text_fragments), None

        return "", f"{extension} files are not supported for textual auto-fill."

    def _infer_text_field(self, filename: str, text: str) -> Optional[str]:
        lowered_name = filename.lower()
        lowered_text = text.lower()

        if 'scope' in lowered_name:
            return 'SCOPE'
        if 'purpose' in lowered_name or 'objective' in lowered_name:
            return 'PURPOSE'

        if lowered_text.count('scope') >= lowered_text.count('purpose'):
            if 'scope' in lowered_text:
                return 'SCOPE'
        if 'purpose' in lowered_text or 'objective' in lowered_text:
            return 'PURPOSE'
        return None

    # ------------------------------------------------------------------ #
    # Helpers
    # ------------------------------------------------------------------ #
    def _build_section_index(self) -> Dict[str, Dict[str, Any]]:
        index: Dict[str, Dict[str, Any]] = {}
        for section in TABLE_CONFIG:
            fields_meta = []
            for field in section.get("fields", []):
                alias_set = self._build_aliases(field)
                fields_meta.append({
                    "ui": field.get("ui"),
                    "doc": field.get("doc"),
                    "form": field.get("form"),
                    "aliases": alias_set,
                })

            min_required = 1 if len(fields_meta) <= 2 else 2
            index[section["ui_section"]] = {
                "fields": fields_meta,
                "min_required": min_required,
            }
        return index

    def _build_aliases(self, field: Dict[str, Any]) -> List[str]:
        aliases = set()
        for key in ("ui", "doc", "form"):
            value = field.get(key)
            if not value:
                continue
            sanitized = value.replace("[]", "")
            aliases.add(self._normalize(sanitized))
        for alias in field.get("aliases", []):
            aliases.add(self._normalize(alias))
        aliases.discard('')
        return list(aliases)

    @staticmethod
    def _normalize(value: Optional[str]) -> str:
        if not value:
            return ''
        return _NORMALIZE.sub('', value.lower())

    @staticmethod
    def _coerce(value: Any) -> str:
        if value is None:
            return ''
        if isinstance(value, str):
            return value.strip()
        return str(value).strip()

    @staticmethod
    def _similarity(a: str, b: str) -> float:
        if not a or not b:
            return 0.0
        if a == b:
            return 1.0
        if a in b or b in a:
            return 0.9
        return SequenceMatcher(None, a, b).ratio()

    @staticmethod
    def _normalize_text_block(value: str) -> str:
        collapsed = re.sub(r'\s+', ' ', value).strip()
        return collapsed

    @staticmethod
    def _clean_row(row: Dict[str, Any]) -> Dict[str, str]:
        return {key: (value or "").strip() for key, value in row.items()}

    @staticmethod
    def _humanize_section(section: str) -> str:
        return section.replace('_', ' ').title()


_parser = SatFormAutoFillParser()


def analyze_sat_upload(file_storage) -> AutoFillResult:
    """Public helper to analyze a FileStorage upload for auto-fill data."""
    try:
        return _parser.process(file_storage)
    except Exception as exc:  # noqa: BLE001
        logger.exception("Auto-fill analysis failed")
        return AutoFillResult(warnings=[f"Auto-fill failed: {exc}"])
