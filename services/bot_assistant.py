import csv
import hashlib
import io
import os
import re
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Set, Tuple

from flask import current_app, session

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - handled gracefully at runtime
    load_workbook = None

try:
    from PIL import Image
except ImportError:  # pragma: no cover - optional dependency
    Image = None

try:
    import requests
except ImportError:  # pragma: no cover - optional dependency
    requests = None

from models import Report, SATReport
from services.ai_assistant import analyze_user_intent
from services.form_autofill import analyze_sat_upload, AutoFillResult


_ALIAS_SANITIZE = re.compile(r'[^a-z0-9]+')
_COLLAPSE_WHITESPACE = re.compile(r'\s+')
EMAIL_PATTERN = re.compile(r'^[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}$', re.IGNORECASE)
NAME_PATTERN = re.compile(r"^[A-Za-z][A-Za-z\s\.\-']+$")
CLIENT_PATTERN = re.compile(r"^[A-Za-z0-9][A-Za-z0-9\s\.\-&']+$")
PROJECT_REFERENCE_PATTERN = re.compile(r'^[A-Z0-9][A-Z0-9_\-./ ]{2,}$')
UUID_PATTERN = re.compile(r'([0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12})', re.IGNORECASE)
SERVICE_OVERVIEW_MESSAGE = (
    "Here are the automation workflows I can help with:\n"
    "- SAT (Site Acceptance Testing) reports\n"
    "- FAT (Factory Acceptance Testing) documentation\n"
    "- Site survey packs and field data collection\n"
    "- Design documentation (FDS, HDS, SDS)\n"
    "- Automation manager reviews\n"
    "- IO builder templates and configurations\n"
    "- Status tracking and compliance reports"
)
GENERAL_KB_RESPONSES = {
    'report_types': SERVICE_OVERVIEW_MESSAGE,
    'progress_help': 'Ask for a "summary" or "what\'s left" whenever you need a checkpoint on collected and pending SAT fields.',
}

SAT_ALIASES = {
    'sat',
    'sat report',
    'site acceptance testing',
    'site acceptance test',
}

NEGATIVE_INTENT_PHRASES = (
    'i do not want',
    "i don't want",
    'i do not need',
    "i don't need",
    'i am not looking',
    'not looking for',
    "don't want",
    "don't need",
    'not interested',
    'something else',
    'different task',
    'no thanks',
    'no thank you',
)
NEGATIVE_INTENT_PREFIXES = (
    'no, i',
    'no i ',
    "no that's",
    'no thats',
    'no, that',
    'no that',
    'no this is not',
    'no this isnt',
)
NEGATIVE_INTENT_COMMANDS = {'cancel', 'stop', 'exit', 'quit', 'no'}

IMAGE_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tif', '.tiff'}
DOCUMENT_EXTENSIONS = {'.txt', '.md', '.docx', '.pdf'}


def _normalize_alias(value: str) -> str:
    if not value:
        return ''
    return _ALIAS_SANITIZE.sub(' ', value.strip().lower()).strip()


def _collapse_whitespace(value: str) -> str:
    return _COLLAPSE_WHITESPACE.sub(' ', value).strip()


def _coerce_to_string(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (datetime, date)):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()
    return str(value).strip()


def _normalize_project_reference(value: str) -> str:
    return _collapse_whitespace(value).upper()


def _normalize_email(value: str) -> str:
    return value.lower()


def _has_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _detect_negative_intent(message: str) -> bool:
    normalized = (message or '').strip().lower()
    if not normalized:
        return False
    if normalized in NEGATIVE_INTENT_COMMANDS:
        return True
    for prefix in NEGATIVE_INTENT_PREFIXES:
        if normalized.startswith(prefix):
            return True
    for phrase in NEGATIVE_INTENT_PHRASES:
        if phrase in normalized:
            return True
    if normalized.startswith('no ') and any(token in normalized for token in (' want', ' need', ' interested', ' looking')):
        return True
    return False


CONVERSATION_FLOW: List[Dict[str, Any]] = [
    {
        "name": "DOCUMENT_TITLE",
        "label": "Document Title",
        "prompt": "Let's start with the SAT document title.",
        "help_text": "Use the exact wording that should appear on the generated report.",
        "required": True,
        "min_length": 3,
        "max_length": 120,
        "aliases": ("document title", "title", "sat title"),
        "normalizer": _collapse_whitespace,
    },
    {
        "name": "CLIENT_NAME",
        "label": "Client Name",
        "prompt": "Who is the client for this report?",
        "help_text": "Enter the organisation or project owner receiving the SAT.",
        "required": True,
        "min_length": 2,
        "max_length": 120,
        "aliases": ("client", "client name", "customer", "company"),
        "normalizer": _collapse_whitespace,
        "pattern": CLIENT_PATTERN,
        "pattern_error": "Client name can include letters, numbers, spaces and -&'.",
    },
    {
        "name": "PROJECT_REFERENCE",
        "label": "Project Reference",
        "prompt": "What is the project reference or identifier?",
        "help_text": "Provide the internal or client reference code used for this SAT.",
        "required": True,
        "min_length": 3,
        "max_length": 60,
        "aliases": ("project reference", "project id", "project", "reference"),
        "normalizer": _normalize_project_reference,
        "pattern": PROJECT_REFERENCE_PATTERN,
        "pattern_error": "Project reference should be at least 3 characters and use letters, numbers or -_/ .",
    },
    {
        "name": "PURPOSE",
        "label": "Purpose",
        "prompt": "Briefly, what is the purpose of this SAT?",
        "help_text": "Summarise why this acceptance test is being conducted.",
        "required": True,
        "min_length": 10,
        "min_words": 3,
        "max_length": 400,
        "aliases": ("purpose", "objective", "aim", "report purpose"),
        "normalizer": _collapse_whitespace,
    },
    {
        "name": "SCOPE",
        "label": "Scope",
        "prompt": "Summarise the scope of the testing.",
        "help_text": "List the major systems or functions covered by the SAT.",
        "required": True,
        "min_length": 10,
        "min_words": 3,
        "max_length": 600,
        "aliases": ("scope", "testing scope", "scope of work"),
        "normalizer": _collapse_whitespace,
    },
]

ADDITIONAL_FIELDS: List[Dict[str, Any]] = [
    {
        "name": "PREPARED_BY",
        "label": "Prepared By",
        "required": False,
        "min_length": 3,
        "max_length": 80,
        "aliases": ("prepared by", "author", "compiled by"),
        "normalizer": _collapse_whitespace,
        "pattern": NAME_PATTERN,
        "pattern_error": "Prepared by should only include alphabetic characters and punctuation.",
    },
    {
        "name": "USER_EMAIL",
        "label": "Prepared By Email",
        "required": False,
        "aliases": ("prepared by email", "author email", "email", "contact email"),
        "pattern": EMAIL_PATTERN,
        "pattern_error": "Please provide a valid email address (e.g. engineer@example.com).",
        "normalizer": _normalize_email,
    },
    {
        "name": "DOCUMENT_REFERENCE",
        "label": "Document Reference",
        "required": False,
        "min_length": 3,
        "max_length": 60,
        "aliases": ("document reference", "doc reference", "reference number", "ref"),
        "normalizer": _normalize_project_reference,
    },
    {
        "name": "REVISION",
        "label": "Revision",
        "required": False,
        "min_length": 1,
        "max_length": 10,
        "aliases": ("revision", "rev"),
        "normalizer": _normalize_project_reference,
    },
]

FIELD_DEFINITIONS: Dict[str, Dict[str, Any]] = {
    item["name"]: dict(item) for item in CONVERSATION_FLOW
}
for item in ADDITIONAL_FIELDS:
    FIELD_DEFINITIONS[item["name"]] = dict(item)

CONVERSATION_ORDER = [item["name"] for item in CONVERSATION_FLOW]
REQUIRED_FIELDS = {name for name, meta in FIELD_DEFINITIONS.items() if meta.get("required")}

_FIELD_ALIAS_LOOKUP: Dict[str, str] = {}
for field_name, meta in FIELD_DEFINITIONS.items():
    for candidate in {
        field_name,
        meta.get("label", ''),
        *meta.get("aliases", ()),
    }:
        normalized = _normalize_alias(candidate)
        if normalized and normalized not in _FIELD_ALIAS_LOOKUP:
            _FIELD_ALIAS_LOOKUP[normalized] = field_name


class BotConversationState:
    """Encapsulates bot state stored in the user session."""

    SESSION_KEY = "bot_conversation_state"

    def __init__(self) -> None:
        self.position = 0
        self.answers: Dict[str, str] = {}
        self.extracted: Dict[str, str] = {}
        self.ingested_files: Dict[str, Dict[str, Any]] = {}
        self.tables: Dict[str, List[Dict[str, Any]]] = {}

    @classmethod
    def load(cls) -> "BotConversationState":
        raw = session.get(cls.SESSION_KEY)
        state = cls()
        if not raw:
            return state
        state.position = raw.get("position", 0)
        state.answers = raw.get("answers", {})
        state.extracted = raw.get("extracted", {})
        state.ingested_files = raw.get("ingested_files", {})
        state.tables = raw.get("tables", {})
        return state

    def save(self) -> None:
        session[self.SESSION_KEY] = {
            "position": self.position,
            "answers": self.answers,
            "extracted": self.extracted,
            "ingested_files": self.ingested_files,
            "tables": self.tables,
        }

    def reset(self) -> None:
        session.pop(self.SESSION_KEY, None)
        self.position = 0
        self.answers = {}
        self.extracted = {}
        self.ingested_files = {}
        self.tables = {}

    def sync_to_next_question(self) -> None:
        """Advance pointer to the next unanswered field."""
        for idx, field_name in enumerate(CONVERSATION_ORDER):
            value = self.answers.get(field_name) or self.extracted.get(field_name)
            if _has_value(value):
                continue
            self.position = idx
            break
        else:
            self.position = len(CONVERSATION_ORDER)


def start_conversation() -> Dict[str, Any]:
    state = BotConversationState()
    state.reset()
    payload = _build_question_payload(state)
    state.save()
    return payload


def process_user_message(message: str, mode: str = "default") -> Dict[str, Any]:
    state = BotConversationState.load()

    if _detect_negative_intent(message):
        state.reset()
        return _build_negative_intent_response()

    general_payload = _handle_general_query(message, state)
    if general_payload is not None:
        state.save()
        return general_payload

    command_payload = _handle_command(message, state)
    if command_payload is not None:
        state.save()
        return command_payload


    primary_intent: Optional[str] = None
    requested_report_type: Optional[str] = None
    treat_as_general = False
    try:
        ai_context = {
            'current_field': CONVERSATION_ORDER[state.position] if state.position < len(CONVERSATION_ORDER) else None,
            'collected': _merge_results(state),
            'pending_fields': _pending_fields(state),
        }
        intent_analysis = analyze_user_intent(message, ai_context)
    except Exception:  # noqa: BLE001 - prefer resilient assistant
        current_app.logger.exception('Failed to analyze user intent for bot assistant.')
        intent_analysis = None

    normalized_report_type: Optional[str] = None
    if intent_analysis:
        primary_intent = intent_analysis.get('intent') or intent_analysis.get('primary_intent')
        confidence = intent_analysis.get('confidence')
        if isinstance(confidence, str):
            try:
                confidence = float(confidence)
            except ValueError:  # noqa: PERF203 - defensive parsing
                confidence = None
        entities = intent_analysis.get('entities') or {}
        for key in ('report_type', 'report', 'document_type', 'workflow'):
            candidate = entities.get(key)
            if isinstance(candidate, dict):
                candidate = candidate.get('name') or candidate.get('value') or candidate.get('type')
            if isinstance(candidate, list):
                candidate = candidate[0] if candidate else None
            if candidate:
                requested_report_type = str(candidate).strip() or None
                break
        if not requested_report_type:
            lowered = message.lower()
            if 'site survey' in lowered:
                requested_report_type = 'site survey'
            elif 'factory acceptance' in lowered or ('fat' in lowered and 'sat' not in lowered):
                requested_report_type = 'fat'
            elif 'fds' in lowered:
                requested_report_type = 'fds'
            elif 'hds' in lowered:
                requested_report_type = 'hds'
            elif 'sds' in lowered:
                requested_report_type = 'sds'
        if requested_report_type:
            normalized_report_type = requested_report_type.lower()
        if primary_intent and primary_intent not in ('create_report', 'workflow_assistance'):
            treat_as_general = True
        elif isinstance(confidence, (int, float)) and confidence is not None and confidence < 0.45:
            treat_as_general = True
        elif primary_intent == 'create_report' and normalized_report_type and normalized_report_type not in SAT_ALIASES:
            treat_as_general = True

    if treat_as_general:
        payload = _handle_general_query(message, state)
        if payload is None:
            payload = _build_question_payload(state)
        metadata = payload.setdefault('metadata', {})
        metadata['agent_intent'] = primary_intent or 'general'
        if requested_report_type:
            metadata['requested_report_type'] = requested_report_type
            friendly = requested_report_type.title()
            payload.setdefault('messages', []).insert(0, f"I can help you start a {friendly} workflow. Open the multi-report AI agent from the toolbar and I'll continue with that report type.")
            suggestions = payload.setdefault('suggestions', [])
            handoff_hint = f"Launch {friendly} assistant"
            if handoff_hint not in suggestions:
                suggestions.append(handoff_hint)
        state.save()
        return payload

    state.sync_to_next_question()
    if state.position >= len(CONVERSATION_ORDER):
        payload = _build_question_payload(state)
        state.save()
        return payload

    field_name = CONVERSATION_ORDER[state.position]
    ok, normalized, error = _apply_validation(field_name, message)
    if not ok:
        meta = FIELD_DEFINITIONS[field_name]
        response: Dict[str, Any] = {
            "completed": False,
            "field": field_name,
            "question": meta["prompt"],
            "collected": _merge_results(state),
            "errors": [error],
            "pending_fields": _pending_fields(state),
        }
        if meta.get("help_text"):
            response["help_text"] = meta["help_text"]
        state.save()
        return response

    if _has_value(normalized):
        state.answers[field_name] = normalized

    state.position += 1
    payload = _build_question_payload(state)
    state.save()
    return payload


def _apply_autofill_updates(state: BotConversationState, autofill_result: Optional[AutoFillResult]):
    """Apply auto-fill results to the current state."""
    applied_fields: Dict[str, str] = {}
    applied_tables: Dict[str, List[Dict[str, str]]] = {}
    validation_warnings: List[str] = []

    if not autofill_result:
        return applied_fields, applied_tables, validation_warnings

    for field_name, raw_value in autofill_result.field_updates.items():
        if field_name not in FIELD_DEFINITIONS:
            continue
        ok, normalized, error = _apply_validation(field_name, raw_value)
        if not ok:
            if error:
                validation_warnings.append(f"{_field_label(field_name)} skipped - {error}")
            continue
        if _has_value(normalized):
            state.extracted[field_name] = normalized
            applied_fields[field_name] = normalized

    for section, rows in autofill_result.table_updates.items():
        normalised_rows = _normalise_table_rows(rows)
        if not normalised_rows:
            continue
        state.tables[section] = normalised_rows
        applied_tables[section] = normalised_rows

    return applied_fields, applied_tables, validation_warnings


def _normalise_table_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    """Ensure table rows are serialisable and non-empty."""
    normalised: List[Dict[str, str]] = []
    for row in rows or []:
        cleaned = {key: _coerce_to_string(value) for key, value in (row or {}).items()}
        if any(_has_value(value) for value in cleaned.values()):
            normalised.append(cleaned)
    return normalised


def _merge_autofill_feedback(
    response: Dict[str, Any],
    applied_fields: Dict[str, str],
    applied_tables: Dict[str, List[Dict[str, str]]],
    autofill_result: Optional[AutoFillResult],
    validation_warnings: List[str],
) -> None:
    """Attach auto-fill metadata to the response payload."""
    if applied_fields:
        response.setdefault("field_updates", {}).update(applied_fields)
    if applied_tables:
        response.setdefault("table_updates", {}).update(applied_tables)

    messages = response.setdefault("messages", [])
    if autofill_result and autofill_result.messages:
        messages.extend(autofill_result.messages)

    warnings: List[str] = []
    if validation_warnings:
        warnings.extend(validation_warnings)
    if autofill_result and autofill_result.warnings:
        warnings.extend(autofill_result.warnings)
    if warnings:
        response.setdefault("warnings", []).extend(warnings)


def ingest_upload(file_storage) -> Dict[str, Any]:
    """Route uploaded assets to the right processor."""
    filename = (file_storage.filename or '').lower()
    extension = os.path.splitext(filename)[1]
    if extension in ('.xlsx', '.xls', '.xlsm'):
        return ingest_excel(file_storage)
    if extension in ('.csv', '.tsv'):
        return ingest_csv(file_storage)
    if extension in DOCUMENT_EXTENSIONS:
        return ingest_document(file_storage)
    if extension in IMAGE_EXTENSIONS:
        return ingest_image(file_storage)
    return _ingest_unknown_file(file_storage)


def ingest_csv(file_storage) -> Dict[str, Any]:
    state = BotConversationState.load()
    try:
        extracted, warnings = _extract_csv_fields(file_storage)
    except Exception as exc:  # noqa: BLE001 - user feedback preferred
        payload = _build_question_payload(state)
        state.save()
        response: Dict[str, Any] = {
            "error": f"Failed to read CSV file: {exc}",
            "collected": payload.get("collected", _merge_results(state)),
            "completed": payload["completed"],
            "pending_fields": payload.get("pending_fields", []),
        }
        if not payload["completed"]:
            response["field"] = payload["field"]
            response["question"] = payload["question"]
            if "help_text" in payload:
                response["help_text"] = payload["help_text"]
        return response

    if extracted:
        state.extracted.update(extracted)

    autofill_result = analyze_sat_upload(file_storage)
    applied_fields, applied_tables, validation_warnings = _apply_autofill_updates(state, autofill_result)

    message = (
        f"Processed {len(extracted)} fields from {file_storage.filename}."
        if extracted
        else f"No recognised fields found in {file_storage.filename}."
    )

    payload = _build_question_payload(state)
    state.save()

    response: Dict[str, Any] = {
        "message": message,
        "collected": payload.get("collected", _merge_results(state)),
        "completed": payload["completed"],
        "pending_fields": payload.get("pending_fields", []),
    }
    response.setdefault("messages", []).append(message)

    if not payload["completed"]:
        response.update({
            "field": payload["field"],
            "question": payload["question"],
        })
        if "help_text" in payload:
            response["help_text"] = payload["help_text"]

    warning_bucket: List[str] = []
    if warnings:
        warning_bucket.extend(warnings)
    _merge_autofill_feedback(response, applied_fields, applied_tables, autofill_result, validation_warnings)
    if warning_bucket:
        response.setdefault("warnings", []).extend(warning_bucket)

    return response


def ingest_document(file_storage) -> Dict[str, Any]:
    state = BotConversationState.load()
    autofill_result = analyze_sat_upload(file_storage)
    applied_fields, applied_tables, validation_warnings = _apply_autofill_updates(state, autofill_result)

    base_message = f"Analysed {file_storage.filename or 'the uploaded document'} for contextual fields."
    payload = _build_question_payload(state)
    state.save()

    response: Dict[str, Any] = {
        "message": base_message,
        "collected": payload.get("collected", _merge_results(state)),
        "completed": payload["completed"],
        "pending_fields": payload.get("pending_fields", []),
    }
    response.setdefault("messages", []).append(base_message)

    if not payload["completed"]:
        response.update({
            "field": payload["field"],
            "question": payload["question"],
        })
        if "help_text" in payload:
            response["help_text"] = payload["help_text"]

    _merge_autofill_feedback(response, applied_fields, applied_tables, autofill_result, validation_warnings)
    return response


def ingest_image(file_storage) -> Dict[str, Any]:
    state = BotConversationState.load()
    filename = file_storage.filename or 'image'
    content = file_storage.read()
    file_storage.stream.seek(0)
    digest = hashlib.sha256(content).hexdigest()

    duplicate = state.ingested_files.get(digest)
    if duplicate:
        payload = _build_question_payload(state)
        state.save()
        warning_text = (
            f"{filename} appears to duplicate {duplicate.get('filename', 'a previous image')}. "
            "I've skipped it to keep the gallery clean."
        )
        payload.setdefault('warnings', []).append(warning_text)
        payload['message'] = warning_text
        return payload

    metadata, analysis_warnings = _analyse_image_stream(content, filename)
    metadata['filename'] = filename
    state.ingested_files[digest] = metadata

    payload = _build_question_payload(state)
    state.save()

    human_resolution = (
        f"{metadata['width']}x{metadata['height']}"
        if metadata.get('width') and metadata.get('height')
        else 'unknown resolution'
    )

    message = (
        f"Stored {filename} ({human_resolution}). I'll stage it for the SAT photo evidence section."
    )

    payload.setdefault('messages', []).append(message)
    payload['message'] = message

    if analysis_warnings:
        payload.setdefault('warnings', []).extend(analysis_warnings)

    payload.setdefault('insights', {})['media'] = list(state.ingested_files.values())

    return payload


def _ingest_unknown_file(file_storage) -> Dict[str, Any]:
    state = BotConversationState.load()
    payload = _build_question_payload(state)
    state.save()
    filename = file_storage.filename or 'the provided file'
    extension = os.path.splitext(filename)[1] or 'unknown format'
    warning_text = (
        f"I can't automate {filename} ({extension}) yet. Try Excel, CSV, or image evidence."
    )
    payload.setdefault('warnings', []).append(warning_text)
    payload['message'] = warning_text
    return payload


def _extract_csv_fields(file_storage) -> Tuple[Dict[str, str], List[str]]:
    content = file_storage.read()
    file_storage.stream.seek(0)
    if not content:
        return {}, ["The CSV file is empty."]

    try:
        text = content.decode('utf-8-sig')
    except UnicodeDecodeError:
        text = content.decode('latin-1', errors='ignore')

    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    extracted: Dict[str, str] = {}
    warnings: List[str] = []

    if not reader.fieldnames:
        warnings.append('Could not detect a header row in the CSV file.')
        return extracted, warnings

    for row_idx, row in enumerate(reader, start=2):
        for header, value in row.items():
            field_name = _match_field_alias(header)
            if not field_name or field_name in extracted:
                continue
            if not _has_value(value):
                continue
            ok, normalized, error = _apply_validation(field_name, value)
            if not ok:
                warnings.append(
                    f"Row {row_idx}: {_field_label(field_name)} skipped - {error}"
                )
                continue
            extracted[field_name] = normalized

    return extracted, warnings


def _analyse_image_stream(content: bytes, filename: str) -> Tuple[Dict[str, Any], List[str]]:
    metadata: Dict[str, Any] = {"size_bytes": len(content)}
    warnings: List[str] = []

    if metadata["size_bytes"] < 10_240:
        warnings.append(
            f"{filename} is very small ({metadata['size_bytes']} bytes); it may not be usable in final reports."
        )

    if Image is None:
        warnings.append('Install Pillow to unlock detailed image analysis.')
        return metadata, warnings

    try:
        with Image.open(io.BytesIO(content)) as img:
            width, height = img.size
            metadata.update({
                'width': width,
                'height': height,
                'format': img.format,
                'mode': img.mode,
            })
            megapixels = (width * height) / 1_000_000
            metadata['megapixels'] = round(megapixels, 2)
            if width * height < 640 * 480:
                warnings.append('The image resolution looks low; consider a higher quality capture.')
    except Exception as exc:  # noqa: BLE001
        warnings.append(f'Unable to analyse {filename}: {exc}')

    return metadata, warnings

def ingest_excel(file_storage) -> Dict[str, Any]:
    state = BotConversationState.load()
    try:
        extracted, warnings = _extract_excel_fields(file_storage)
    except Exception as exc:  # noqa: BLE001 - user facing feedback preferred here
        payload = _build_question_payload(state)
        state.save()

        response: Dict[str, Any] = {
            "error": f"Failed to read Excel file: {exc}",
            "collected": payload.get("collected", _merge_results(state)),
            "completed": payload["completed"],
            "pending_fields": payload.get("pending_fields", []),
        }

        if not payload["completed"]:
            response["field"] = payload["field"]
            response["question"] = payload["question"]
            if "help_text" in payload:
                response["help_text"] = payload["help_text"]

        return response

    if extracted:
        state.extracted.update(extracted)

    autofill_result = analyze_sat_upload(file_storage)
    applied_fields, applied_tables, validation_warnings = _apply_autofill_updates(state, autofill_result)

    message = (
        f"Processed {len(extracted)} fields from {file_storage.filename}."
        if extracted
        else f"No recognised fields found in {file_storage.filename}."
    )

    payload = _build_question_payload(state)
    state.save()

    response: Dict[str, Any] = {
        "message": message,
        "collected": payload.get("collected", _merge_results(state)),
        "completed": payload["completed"],
        "pending_fields": payload.get("pending_fields", []),
    }
    response.setdefault("messages", []).append(message)

    if not payload["completed"]:
        response.update({
            "field": payload["field"],
            "question": payload["question"],
        })
        if "help_text" in payload:
            response["help_text"] = payload["help_text"]

    warning_bucket: List[str] = []
    if warnings:
        warning_bucket.extend(warnings)

    _merge_autofill_feedback(response, applied_fields, applied_tables, autofill_result, validation_warnings)

    if warning_bucket:
        response.setdefault("warnings", []).extend(warning_bucket)

    return response


def _extract_excel_fields(file_storage) -> Tuple[Dict[str, str], List[str]]:
    if load_workbook is None:
        raise RuntimeError('Excel support requires the openpyxl package to be installed on the server.')

    content = file_storage.read()
    workbook = load_workbook(filename=io.BytesIO(content), data_only=True)
    file_storage.stream.seek(0)

    extracted: Dict[str, str] = {}
    warnings: List[str] = []

    for sheet in workbook.worksheets:
        sheet_data, sheet_warnings = _extract_fields_from_sheet(sheet)
        for field_name, value in sheet_data.items():
            if field_name in extracted:
                warnings.append(f"{sheet.title}: duplicate value for {_field_label(field_name)} ignored.")
                continue
            extracted[field_name] = value
        warnings.extend(sheet_warnings)

    return extracted, warnings


def _extract_fields_from_sheet(sheet) -> Tuple[Dict[str, str], List[str]]:
    results: Dict[str, str] = {}
    warnings: List[str] = []

    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if not row:
            continue
        for col_idx, cell in enumerate(row):
            header = _coerce_to_string(cell)
            if not header:
                continue
            field_name = _match_field_alias(header)
            if not field_name or field_name in results:
                continue

            value = _extract_candidate_value(sheet, row_idx, col_idx, row)
            if value is None:
                continue

            ok, normalized, error = _apply_validation(field_name, value)
            if not ok:
                warnings.append(
                    f"{sheet.title}: {_field_label(field_name)} at row {row_idx} skipped - {error}"
                )
                continue

            if _has_value(normalized):
                results[field_name] = normalized

    return results, warnings


def _extract_candidate_value(sheet, row_idx: int, col_idx: int, row) -> Optional[str]:
    # Prefer same-row values, fall back to a short lookahead in the same column.
    for candidate in row[col_idx + 1:]:
        candidate_value = _coerce_to_string(candidate)
        if candidate_value:
            return candidate_value

    max_lookahead = 3
    for offset in range(1, max_lookahead + 1):
        target_row = row_idx + offset
        if target_row > sheet.max_row:
            break
        candidate_value = _coerce_to_string(sheet.cell(row=target_row, column=col_idx + 1).value)
        if candidate_value:
            return candidate_value

    return None


def _match_field_alias(header: str) -> Optional[str]:
    normalized = _normalize_alias(header)
    if not normalized:
        return None
    return _FIELD_ALIAS_LOOKUP.get(normalized)


def _field_label(field_name: str) -> str:
    meta = FIELD_DEFINITIONS.get(field_name, {})
    return meta.get("label", field_name.replace('_', ' ').title())


def _apply_validation(field_name: str, raw_value: Any) -> Tuple[bool, Optional[str], Optional[str]]:
    meta = FIELD_DEFINITIONS[field_name]
    label = meta.get("label", field_name.replace('_', ' ').title())
    value = _coerce_to_string(raw_value)

    if not value:
        if meta.get("required"):
            return False, None, f"{label} is required."
        return True, "", None

    if meta.get("strip", True):
        value = value.strip()

    min_length = meta.get("min_length")
    if min_length and len(value) < min_length:
        return False, None, f"{label} must be at least {min_length} characters long."

    min_words = meta.get("min_words")
    if min_words and len(value.split()) < min_words:
        return False, None, f"{label} should contain at least {min_words} words."

    max_length = meta.get("max_length")
    if max_length and len(value) > max_length:
        value = value[:max_length].strip()

    pattern = meta.get("pattern")
    if pattern and not pattern.fullmatch(value):
        return False, None, meta.get("pattern_error", f"{label} has an invalid format.")

    normalizer = meta.get("normalizer")
    if normalizer:
        value = normalizer(value)

    validator = meta.get("validator")
    if validator:
        try:
            value = validator(value, meta)
        except ValueError as exc:
            return False, None, str(exc)

    return True, value, None


def _merge_results(state: BotConversationState) -> Dict[str, str]:
    merged: Dict[str, str] = {}
    combined = dict(state.extracted)
    combined.update(state.answers)
    for key, value in combined.items():
        if _has_value(value):
            merged[key] = value
    return merged


def _pending_fields(state: BotConversationState) -> List[str]:
    pending: List[str] = []
    merged = _merge_results(state)
    for field_name in CONVERSATION_ORDER:
        if not _has_value(merged.get(field_name)):
            pending.append(field_name)
    return pending


def _build_negative_intent_response() -> Dict[str, Any]:
    return {
        'completed': False,
        'collected': {},
        'pending_fields': [],
        'messages': [
            "No problem! I understand you don't need that right now.",
            SERVICE_OVERVIEW_MESSAGE,
            "What would you like to work on instead?",
        ],
        'question': 'Is there another workflow I can help you with?',
    }


def _build_question_payload(state: BotConversationState) -> Dict[str, Any]:
    state.sync_to_next_question()
    merged = _merge_results(state)
    pending = _pending_fields(state)

    if state.position >= len(CONVERSATION_ORDER):
        return {
            "completed": True,
            "collected": merged,
            "pending_fields": pending,
        }

    field_name = CONVERSATION_ORDER[state.position]
    meta = FIELD_DEFINITIONS[field_name]
    payload: Dict[str, Any] = {
        "completed": False,
        "field": field_name,
        "question": meta["prompt"],
        "collected": merged,
        "pending_fields": pending,
    }

    if meta.get("help_text"):
        payload["help_text"] = meta["help_text"]

    return payload


def _parse_document_request(message: str) -> Optional[str]:
    lowered = message.lower().strip()
    if not lowered:
        return None
    if not any(keyword in lowered for keyword in ("download", "fetch", "get")):
        return None
    match = UUID_PATTERN.search(message)
    if match:
        return match.group(1)
    return None


def _parse_summary_request(message: str) -> bool:
    lowered = message.lower().strip()
    if not lowered:
        return False
    triggers = ("summary", "status", "progress", "what's left", "what is left")
    return any(trigger in lowered for trigger in triggers)


def _format_summary_message(summary: Dict[str, Any]) -> str:
    collected = summary.get('collected') or {}
    pending = summary.get('pending_fields') or []
    collected_count = len(collected)
    if not pending:
        return f"All required SAT fields are captured ({collected_count} fields). You're ready to continue."
    pending_labels = [_field_label(name) for name in pending]
    pending_text = _format_human_list(pending_labels)
    return f"Collected {collected_count} fields so far. Still waiting on {pending_text}."


def _format_human_list(items: List[str]) -> str:
    if not items:
        return ''
    if len(items) == 1:
        return items[0]
    if len(items) == 2:
        return f"{items[0]} and {items[1]}"
    return ', '.join(items[:-1]) + f", and {items[-1]}"



def _handle_research_request(query: str, state: BotConversationState) -> Dict[str, Any]:
    state.sync_to_next_question()
    payload = _build_question_payload(state)
    payload.setdefault("messages", [])
    context = _merge_results(state)
    insights, warnings = _perform_external_research(query, context)

    normalized_query = _collapse_whitespace(query)
    if normalized_query:
        payload["messages"].append(
            f'Analysed live sources for "{normalized_query}".'
        )
    else:
        warnings.append('Provide a topic or question for me to research.')

    if insights:
        for insight in insights:
            title = insight.get('title') or 'Insight'
            summary = insight.get('summary') or ''
            line = f"{title}: {summary}".strip()
            if insight.get('url'):
                line += f" ({insight['url']})"
            payload["messages"].append(line)
        payload['research'] = insights
    elif not warnings:
        payload["messages"].append(
            "No external intelligence surfaced yet, but I'll keep leveraging internal context."
        )

    if warnings:
        payload.setdefault('warnings', []).extend(warnings)

    return payload


def _perform_external_research(query: str, context: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], List[str]]:
    warnings: List[str] = []
    clean_query = _collapse_whitespace(query)
    if not clean_query:
        return [], ['I need a topic to research.']

    if requests is None:
        warnings.append('External research requires the requests package.')
        return [], warnings

    app = current_app._get_current_object()
    if not app.config.get('ASSISTANT_ALLOW_EXTERNAL', True):
        warnings.append('External research is disabled for this environment.')
        return [], warnings

    search_terms = [clean_query]
    for key in ("PROJECT_REFERENCE", "CLIENT_NAME", "PURPOSE"):
        value = context.get(key)
        if value:
            search_terms.append(str(value))
    search_query = ' '.join(search_terms)

    params = {
        'q': search_query,
        'format': 'json',
        'no_html': 1,
        'no_redirect': 1,
        'skip_disambig': 1,
    }

    try:
        timeout = app.config.get('ASSISTANT_RESEARCH_TIMEOUT', 8)
        response = requests.get('https://api.duckduckgo.com/', params=params, timeout=timeout)
        response.raise_for_status()
        data = response.json()
    except requests.RequestException as exc:  # pragma: no cover - best effort outreach
        warnings.append(f'External research failed: {exc}')
        return [], warnings

    insights: List[Dict[str, Any]] = []

    abstract = (data.get('AbstractText') or '').strip()
    if abstract:
        insights.append({
            'title': data.get('Heading') or 'Instant insight',
            'summary': abstract,
            'url': data.get('AbstractURL'),
        })

    for topic in data.get('RelatedTopics', []):
        insights.extend(_parse_duckduckgo_topic(topic))
        if len(insights) >= 5:
            break

    filtered: List[Dict[str, Any]] = []
    seen_titles: Set[str] = set()
    for insight in insights:
        title = insight.get('title') or 'Insight'
        if title in seen_titles:
            continue
        summary = insight.get('summary') or ''
        if len(summary) > 240:
            summary = summary[:237].rstrip() + '...'
        filtered.append({
            'title': title,
            'summary': summary,
            'url': insight.get('url'),
        })
        seen_titles.add(title)
        if len(filtered) >= 3:
            break

    if not filtered:
        warnings.append('No external intelligence returned yet for that request.')

    return filtered, warnings


def _parse_duckduckgo_topic(topic: Any) -> List[Dict[str, Any]]:
    results: List[Dict[str, Any]] = []
    if not isinstance(topic, dict):
        return results
    text = topic.get('Text')
    if text:
        title = text.split(' - ')[0]
        results.append({
            'title': title,
            'summary': text,
            'url': topic.get('FirstURL'),
        })
    for nested in topic.get('Topics', []) or []:
        results.extend(_parse_duckduckgo_topic(nested))
    return results



def _handle_general_query(message: str, state: BotConversationState) -> Optional[Dict[str, Any]]:
    normalized = message.strip().lower()
    if not normalized:
        return None

    responses: List[str] = []

    service_requested = False
    if (
        'document type' in normalized
        or 'report type' in normalized
        or 'report template' in normalized
        or ('report' in normalized and any(token in normalized for token in ('available', 'options', 'list', 'types')))
        or ('services' in normalized and any(token in normalized for token in ('offer', 'available', 'options', 'provide', 'list')))
        or ('what' in normalized and 'services' in normalized)
        or 'what can you do' in normalized
        or 'what all can you do' in normalized
        or 'what do you do' in normalized
        or 'what are you capable' in normalized
        or 'what are your capabilities' in normalized
    ):
        responses.append(SERVICE_OVERVIEW_MESSAGE)
        service_requested = True

    if ('summary' in normalized or "what's left" in normalized) and not service_requested:
        responses.append(GENERAL_KB_RESPONSES['progress_help'])

    if 'help' in normalized and 'sat' in normalized:
        responses.append('I can guide you through each SAT field, fill values from spreadsheets, and fetch generated documents on demand.')

    if not responses:
        return None

    if service_requested:
        return {
            'completed': False,
            'collected': _merge_results(state),
            'pending_fields': _pending_fields(state),
            'messages': responses,
            'question': 'Would you like me to start one of these workflows?',
        }

    payload = _build_question_payload(state)
    payload.setdefault('messages', []).extend(responses)
    return payload

def _handle_command(message: str, state: BotConversationState) -> Optional[Dict[str, Any]]:
    submission_id = _parse_document_request(message)
    if submission_id:
        result = resolve_report_download_url(submission_id)
        payload = _build_question_payload(state)

        command: Dict[str, Any] = {
            'type': 'document_fetch',
            'requested_id': submission_id,
            'success': 'download_url' in result,
        }

        if command['success']:
            command['download_url'] = result['download_url']
            metadata = {
                key: result[key]
                for key in ('document_title', 'client_name', 'project_reference')
                if result.get(key)
            }
            if metadata:
                command['metadata'] = metadata
            title = metadata.get('document_title') if metadata else None
            if not title:
                title = 'SAT Report'
            payload.setdefault('messages', []).append(
                f"{title} is ready to download."
            )
        else:
            command['error'] = result.get('error', 'Document not found.')
            payload.setdefault('command_errors', []).append(command['error'])
            payload.setdefault('messages', []).append(command['error'])

        payload['command'] = command
        return payload

    if _parse_summary_request(message):
        payload = _build_question_payload(state)
        collected = _merge_results(state)
        summary_command: Dict[str, Any] = {
            'type': 'summary',
            'completed': payload['completed'],
            'collected': collected,
            'pending_fields': payload.get('pending_fields', []),
            'missing_required': [
                field for field in REQUIRED_FIELDS if not _has_value(collected.get(field))
            ],
        }
        payload['command'] = summary_command
        payload.setdefault('messages', []).append(
            _format_summary_message(summary_command)
        )
        return payload

    return None
def resolve_report_download_url(submission_id: str) -> Dict[str, str]:
    report = Report.query.filter_by(id=submission_id).first()
    if not report:
        return {"error": "Report not found."}
    sat_report = SATReport.query.filter_by(report_id=submission_id).first()
    if not sat_report:
        return {"error": "SAT data not available for this report."}
    return {
        "download_url": f"/status/download-modern/{submission_id}",
        "document_title": report.document_title or "SAT Report",
        "client_name": report.client_name,
        "project_reference": report.project_reference,
    }
